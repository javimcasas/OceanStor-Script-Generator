import tkinter as tk
from tkinter import messagebox
import subprocess
import os, sys
from utils import load_config, get_data_file_path
from file_operations import create_excel_for_resource
from gui_helpers import apply_style, create_buttons_and_dropdown
from command_generator import main as generate_commands

def run_script(script_type, command_type):
    try:
        config = load_config()
        if script_type not in config:
            raise ValueError(f"Invalid script type: {script_type}. Valid types are: {list(config.keys())}")

        documents_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Documents")
        excel_path = os.path.join(documents_folder, f'{script_type}_commands.xlsx')

        if not os.path.exists(excel_path):
            messagebox.showinfo("Info", f"Please create the Excel file first by pressing the 'Create Excel' button.")
            return

        # Call the command generator logic directly
        output_file_path = generate_commands(script_type)

        if os.path.exists(output_file_path):
            messagebox.showinfo("Success", f"Script executed successfully. Output saved to:\n{output_file_path}")
            os.startfile(output_file_path)  # Open the output file
        else:
            messagebox.showerror("Error", f"The file '{output_file_path}' does not exist.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

def open_excel_with_sheet(excel_path, sheet_name):
    try:
        from openpyxl import load_workbook
        try:
            # Attempt to open the workbook
            workbook = load_workbook(excel_path)
            if sheet_name in workbook.sheetnames:
                workbook.active = workbook[sheet_name]
                workbook.save(excel_path)
            os.startfile(excel_path)
        except PermissionError:
            # If the file is already open, inform the user
            messagebox.showinfo("Info", f"The Excel file '{excel_path}' is already open. Please close it and try again.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while opening the Excel sheet: {e}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while opening the Excel sheet: {e}")

def open_excel(script_type, command_type):
    try:
        config = load_config()
        if script_type not in config:
            raise ValueError(f"Invalid script type: {script_type}. Valid types are: {list(config.keys())}")

        documents_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Documents")
        if not os.path.exists(documents_folder):
            os.makedirs(documents_folder)

        excel_path = os.path.join(documents_folder, f'{script_type}_commands.xlsx')
        if not os.path.exists(excel_path):
            create_excel_for_resource(script_type, excel_path)

        if os.path.exists(excel_path):
            open_excel_with_sheet(excel_path, command_type)
        else:
            messagebox.showerror("Error", f"The file '{excel_path}' does not exist.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while trying to open the Excel file: {e}")

def clear_excel(script_type):
    try:
        config = load_config()
        if script_type not in config:
            raise ValueError(f"Invalid script type: {script_type}. Valid types are: {list(config.keys())}")

        documents_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Documents")
        excel_path = os.path.join(documents_folder, f'{script_type}_commands.xlsx')

        if os.path.exists(excel_path):
            confirm = messagebox.askyesno("Confirm Deletion", f"Are you sure you want to delete the Excel file for {script_type}?")
            if confirm:
                os.remove(excel_path)
                messagebox.showinfo("Success", f"The Excel file for {script_type} has been deleted.")
        else:
            messagebox.showinfo("Info", f"No Excel file found for {script_type}.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while deleting the Excel file: {e}")

def main():
    root = tk.Tk()
    root.title("Script Selector")

    window_width = 550
    window_height = 450
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    position_x = (screen_width - window_width) // 2
    position_y = (screen_height - window_height) // 2
    root.geometry(f"{window_width}x{window_height}+{position_x}+{position_y}")

    apply_style(root)

    script_var = tk.StringVar(value="CIFS")
    command_var = tk.StringVar(value="Create")

    config = load_config()
    script_menu, open_button, run_button, clear_button = create_buttons_and_dropdown(
        root, open_excel, run_script, script_var, command_var, clear_excel, config
    )

    root.mainloop()

if __name__ == "__main__":
    main()