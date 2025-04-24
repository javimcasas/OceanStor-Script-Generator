import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from utils import load_config

def create_excel_for_resource(resource_type, excel_path, device_var, num_rows=1000):
    """Create an Excel file for a specific resource type with styling and multiple rows for user input."""
    if not os.path.exists(excel_path):
        config = load_config(device_var)
        resource_config = config.get(resource_type, {}).get('operations', {})

        # Create a new workbook
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove the default sheet

        # Define styles
        mandatory_fill = PatternFill(start_color="FFD3D3", end_color="FFD3D3", fill_type="solid")  # Light red for mandatory fields
        optional_fill = PatternFill(start_color="D3FFD3", end_color="D3FFD3", fill_type="solid")  # Light green for optional fields
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue for headers
        header_font = Font(color="FFFFFF", bold=True)  # White and bold for headers
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        alignment = Alignment(horizontal="center", vertical="center")

        # Create sheets for each command type (Create, Change, Show, etc.)
        for command_type, command_config in resource_config.items():
            sheet = workbook.create_sheet(title=command_type)
            
            # Combine mandatory and optional fields
            all_fields = command_config.get('mandatory', []) + command_config.get('optional', [])
            
            # Write headers
            headers = [field["name"] for field in all_fields]
            sheet.append(headers)

            # Apply header styles
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = alignment

            # Add data validation and apply styles
            for col_idx, field in enumerate(all_fields, start=1):
                field_type = field.get('field_type', 'text')
                
                # Add dropdown for select fields
                if field_type == "select":
                    dv = DataValidation(
                        type="list",
                        formula1=f'"{",".join(field["allowed_values"])}"',
                        allow_blank=True
                    )
                    dv.error = "Invalid value. Please select from the dropdown."
                    dv.errorTitle = "Invalid Entry"
                    dv.prompt = "Please select a value from the dropdown."
                    dv.promptTitle = "Select Value"
                    sheet.add_data_validation(dv)
                    
                    column_letter = get_column_letter(col_idx)
                    dv.add(f"{column_letter}2:{column_letter}{num_rows + 1}")

                # Apply cell styles
                is_mandatory = field in command_config.get('mandatory', [])
                for row_idx in range(2, num_rows + 2):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.fill = mandatory_fill if is_mandatory else optional_fill
                    cell.border = border
                    cell.alignment = alignment

            # Adjust column widths
            for col_idx, header in enumerate(headers, start=1):
                column_letter = get_column_letter(col_idx)
                sheet.column_dimensions[column_letter].width = max(len(header) + 2, 15)

        # Save the workbook
        workbook.save(excel_path)
        print(f"Excel template created for {resource_type} at: {excel_path}")
        return True
    else:
        print(f"Excel file already exists: {excel_path}")
        return False