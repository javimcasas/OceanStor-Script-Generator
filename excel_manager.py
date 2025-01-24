import os
import json
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter  # Import utility for column letter conversion

def load_config():
    """Load the configuration from the JSON file."""
    with open("commands_config.json", "r") as file:
        return json.load(file)

def create_excel_for_resource(resource_type, excel_path, output_folder):
    """
    Creates an Excel file for a specific resource type (e.g., CIFS, NFS, FileSystem).
    Each sheet in the Excel file corresponds to a command type (Create, Change, Show).

    :param resource_type: The resource type (e.g., "CIFS", "NFS", "FileSystem").
    :param excel_path: The full path to the Excel file.
    :param output_folder: The folder where the Excel file will be saved.
    """
    if not os.path.exists(excel_path):
        # Load configuration from the JSON file
        config = load_config()
        resource_config = config.get(resource_type, {})

        # Create a new workbook
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove the default sheet

        # Create sheets for each command type (Create, Change, Show)
        for command_type, fields in resource_config.items():
            # Create a new sheet
            sheet = workbook.create_sheet(title=command_type)

            # Write headers
            headers = [field["name"] for field in fields.get("mandatory", []) + fields.get("optional", [])]
            sheet.append(headers)

            # Add data validation for selectable fields
            for col_idx, field in enumerate(fields.get("mandatory", []) + fields.get("optional", []), start=1):
                if field.get("field_type") == "select":
                    # Create a data validation object
                    dv = DataValidation(
                        type="list",
                        formula1=f'"{",".join(field["allowed_values"])}"',
                        allow_blank=True
                    )
                    dv.error = "Invalid value. Please select from the dropdown."
                    dv.errorTitle = "Invalid Entry"
                    dv.prompt = "Please select a value from the dropdown."
                    dv.promptTitle = "Select Value"
                    sheet.add_data_validation(dv)

                    # Apply data validation to the first 1000 rows of the column
                    column_letter = get_column_letter(col_idx)  # Convert column index to letter (e.g., 1 -> 'A')
                    dv.add(f"{column_letter}2:{column_letter}1000")

        # Save the workbook
        workbook.save(excel_path)
        print(f"Excel file created for {resource_type}: {excel_path}")
    else:
        print(f"Excel file already exists: {excel_path}")