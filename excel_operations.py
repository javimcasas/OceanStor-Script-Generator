import os
from tkinter import messagebox, filedialog
from utils import load_config, encapsulate_results, open_directory
from file_operations import create_excel_for_resource
from import_operations import process_imported_data
from gui_functions import toggle_loading

def open_excel_with_sheet(excel_path, sheet_name):
    try:
        from openpyxl import load_workbook
        try:
            # Attempt to open the workbook
            workbook = load_workbook(excel_path)
            if sheet_name in workbook.sheetnames:
                workbook.active = workbook[sheet_name]
                workbook.save(excel_path)
            os.startfile(excel_path)
        except PermissionError:
            # If the file is already open, inform the user
            messagebox.showinfo("Info", f"The Excel file '{excel_path}' is already open. Please close it and try again.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while opening the Excel sheet: {e}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while opening the Excel sheet: {e}")

def open_excel(script_type, command_type, device_var):
    try:
        config = load_config(device_var)
        if script_type not in config:
            raise ValueError(f"Invalid script type: {script_type}. Valid types are: {list(config.keys())}")

        documents_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Documents")
        if not os.path.exists(documents_folder):
            os.makedirs(documents_folder)

        excel_path = os.path.join(documents_folder, f'{device_var}_{script_type}_commands.xlsx')
        if not os.path.exists(excel_path):
            create_excel_for_resource(script_type, excel_path, device_var)

        if os.path.exists(excel_path):
            open_excel_with_sheet(excel_path, command_type)
        else:
            messagebox.showerror("Error", f"The file '{excel_path}' does not exist.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while trying to open the Excel file: {e}")

def clear_excel(script_type, device_var):
    try:
        config = load_config(device_var)
        if script_type not in config:
            raise ValueError(f"Invalid script type: {script_type}. Valid types are: {list(config.keys())}")

        documents_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Documents")
        excel_path = os.path.join(documents_folder, f'{device_var}_{script_type}_commands.xlsx')

        if os.path.exists(excel_path):
            confirm = messagebox.askyesno("Confirm Deletion", f"Are you sure you want to delete the Excel file for {script_type}?")
            if confirm:
                os.remove(excel_path)
                messagebox.showinfo("Success", f"The Excel file for {script_type} has been deleted.")
        else:
            messagebox.showinfo("Info", f"No Excel file found for {script_type}.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while deleting the Excel file: {e}")
        
def import_excel(root):
    """Handle Excel file import with proper loading visibility."""
    try:
        # Open file browser for Excel files
        file_path = filedialog.askopenfilename(
            title="Select Excel File to Import",
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")]
        )
        
        if not file_path:  # User cancelled
            return

        # Show loading (force UI update)
        toggle_loading(root, True, "Processing Excel file...")
        root.update()  # Force immediate UI refresh (stronger than update_idletasks)
        
        try:
            # Process the file
            process_imported_data(file_path)
            
            # Update loading message (force UI update)
            toggle_loading(root, True, "Creating zip archive...")
            root.update()
            
            zip_path = encapsulate_results()
            
            # Success message
            success_msg = "Excel file processed successfully!"
            if zip_path:
                success_msg += f"\n\nZipped as: {os.path.basename(zip_path)}"
            
            # Hide loading before showing message box
            toggle_loading(root, False)
            root.update()  # Ensure loading is gone before dialog
            
            # Ask to open folder
            response = messagebox.askyesno(
                "Success", 
                f"{success_msg}\n\nOpen Imported_Results folder?"
            )
            
            if response:
                open_directory("Imported_Results")
                
        except Exception as e:
            toggle_loading(root, False)
            root.update()
            raise  # Re-raise to outer handler
            
    except Exception as e:
        toggle_loading(root, False)
        messagebox.showerror("Error", f"Error importing Excel: {str(e)}")