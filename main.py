import tkinter as tk
from tkinter import messagebox, ttk, filedialog
import subprocess
import os, sys
from utils import load_config, get_data_file_path, open_directory, encapsulate_results
from file_operations import create_excel_for_resource
from gui_helpers import apply_style, create_buttons_and_dropdown, toggle_loading
from command_generator import main as generate_commands
from import_operations import process_imported_data

def run_script(script_type, command_type, device_type):
    try:
        config = load_config(device_type)
        if script_type not in config:
            raise ValueError(f"Invalid script type: {script_type}. Valid types are: {list(config.keys())}")

        documents_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Documents")
        excel_path = os.path.join(documents_folder, f'{device_type}_{script_type}_commands.xlsx')

        if not os.path.exists(excel_path):
            messagebox.showinfo("Info", f"Please create the Excel file first by pressing the 'Create Excel' button.")
            return

        # Call the command generator logic directly
        output_file_path = generate_commands(script_type, device_type)

        if os.path.exists(output_file_path):
            messagebox.showinfo("Success", f"Script executed successfully. Output saved to:\n{output_file_path}")
            os.startfile(output_file_path)  # Open the output file
        else:
            messagebox.showerror("Error", f"The file '{output_file_path}' does not exist.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

def open_excel_with_sheet(excel_path, sheet_name):
    try:
        from openpyxl import load_workbook
        try:
            # Attempt to open the workbook
            workbook = load_workbook(excel_path)
            if sheet_name in workbook.sheetnames:
                workbook.active = workbook[sheet_name]
                workbook.save(excel_path)
            os.startfile(excel_path)
        except PermissionError:
            # If the file is already open, inform the user
            messagebox.showinfo("Info", f"The Excel file '{excel_path}' is already open. Please close it and try again.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while opening the Excel sheet: {e}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while opening the Excel sheet: {e}")

def open_excel(script_type, command_type, device_var):
    try:
        config = load_config(device_var)
        if script_type not in config:
            raise ValueError(f"Invalid script type: {script_type}. Valid types are: {list(config.keys())}")

        documents_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Documents")
        if not os.path.exists(documents_folder):
            os.makedirs(documents_folder)

        excel_path = os.path.join(documents_folder, f'{device_var}_{script_type}_commands.xlsx')
        if not os.path.exists(excel_path):
            create_excel_for_resource(script_type, excel_path, device_var)

        if os.path.exists(excel_path):
            open_excel_with_sheet(excel_path, command_type)
        else:
            messagebox.showerror("Error", f"The file '{excel_path}' does not exist.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while trying to open the Excel file: {e}")

def clear_excel(script_type, device_var):
    try:
        config = load_config(device_var)
        if script_type not in config:
            raise ValueError(f"Invalid script type: {script_type}. Valid types are: {list(config.keys())}")

        documents_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Documents")
        excel_path = os.path.join(documents_folder, f'{device_var}_{script_type}_commands.xlsx')

        if os.path.exists(excel_path):
            confirm = messagebox.askyesno("Confirm Deletion", f"Are you sure you want to delete the Excel file for {script_type}?")
            if confirm:
                os.remove(excel_path)
                messagebox.showinfo("Success", f"The Excel file for {script_type} has been deleted.")
        else:
            messagebox.showinfo("Info", f"No Excel file found for {script_type}.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while deleting the Excel file: {e}")
        
def import_excel(root):
    """Handle Excel file import functionality."""
    try:
        # Open file browser for Excel files
        file_path = filedialog.askopenfilename(
            title="Select Excel File to Import",
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")]
        )
        
        if not file_path:  # User cancelled
            return

        # Show loading indicator
        toggle_loading(root, True, "Processing Excel file...")
        
        try:
            # Process the file directly
            process_imported_data(file_path)
            
            # Encapsulate results into zip file
            zip_path = encapsulate_results()
            
            # Prepare success message
            success_msg = "Excel file processed successfully!"
            if zip_path:
                success_msg += f"\n\nAll command files have been zipped as:\n{os.path.basename(zip_path)}"
            
            # Ask user if they want to open the results folder
            response = messagebox.askyesno(
                "Success", 
                f"{success_msg}\n\nWould you like to open the Imported_Results folder?"
            )
            
            if response:
                open_directory("Imported_Results")
        finally:
            # Always hide loading indicator when done
            toggle_loading(root, False)

    except Exception as e:
        toggle_loading(root, False)  # Ensure loading is hidden on error
        messagebox.showerror("Error", f"An error occurred while importing Excel: {str(e)}")

def main():
    root = tk.Tk()
    root.title("Script Selector")
    
    try:
        icon_path = os.path.join("Icons", "exe_icon.ico")
        root.iconbitmap(icon_path)
    except Exception as e:
        try:
            base_path = sys._MEIPASS if getattr(sys, 'frozen', False) else os.path.abspath(".")
            icon_path = os.path.join(base_path, "Icons", "exe_icon.ico")
            root.iconbitmap(icon_path)
        except Exception as e:
            print(f"Could not load window icon: {str(e)}")

    window_width = 550
    window_height = 450
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    position_x = (screen_width - window_width) // 2
    position_y = (screen_height - window_height) // 2
    root.geometry(f"{window_width}x{window_height}+{position_x}+{position_y}")

    apply_style(root)

    # Add a new dropdown for device selection (OceanStor Dorado or OceanStor Pacific)
    device_var = tk.StringVar(value="OceanStor Dorado")  # Default to OceanStor Dorado
    script_var = tk.StringVar(value="CIFS")  # Default script type
    command_var = tk.StringVar(value="Create")  # Default command type

    # Device selection dropdown
    device_menu = ttk.Combobox(
        root,
        textvariable=device_var,
        values=["OceanStor Dorado", "OceanStor Pacific"],  # Options for device selection
        state="readonly",
        font=("Arial", 11),
        justify="center",
    )
    device_menu.pack(pady=10)

    # Load config based on the selected device
    def load_selected_config():
        selected_device = device_var.get()
        return load_config(selected_device)

    # Create buttons and dropdowns for script and command selection
    config = load_selected_config()
    script_menu, open_button, run_button, clear_button, update_command_buttons = create_buttons_and_dropdown(
        root, open_excel, run_script, script_var, command_var, clear_excel, import_excel, device_var, config
    )

    # Update the script dropdown and buttons when the device changes
    def update_script_dropdown(*args):
        selected_device = device_var.get()
        print(f"Selected device: {selected_device}")  # Debug print
        new_config = load_config(selected_device)  

        if script_var.get() not in new_config:
            script_var.set(list(new_config.keys())[0])  # Reset only if needed
        
        script_menu['values'] = list(new_config.keys())
        update_command_buttons(new_config)

    device_var.trace_add("write", update_script_dropdown)

    # Update the buttons when the script type changes
    def update_buttons(*args):
        selected_script = script_var.get()
        selected_device = device_var.get()
        new_config = load_config(selected_device)

        if selected_script in new_config:
            update_command_buttons(new_config)

    script_var.trace_add("write", update_buttons)

    root.mainloop()

if __name__ == "__main__":
    main()